import streamlit as st
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Font
from io import BytesIO

def merge_inv_rpt(Inv_List):
    # report workbook before being saved
    Rpt = Workbook()
    Rpt.remove(Rpt.active)
    # read all tank inventory excel files
    for Inv_file in Inv_List:
        # open each product inventory workbook
        Inv = openpyxl.load_workbook(Inv_file)
        for sheet in Inv.sheetnames:
             # open each product inventory sheet
            Inv_sheet = Inv[sheet]
             # copy each product inventory sheet to the report workbook
            Rpt_sheet = Rpt.create_sheet(Inv_file.name.replace('.xlsx', ''))
            for row in Inv_sheet.iter_rows(values_only=True):
                Rpt_sheet.append(row)
            # apply borders to the report workbook
            for row in Rpt_sheet.iter_rows(min_row=1, max_row=Rpt_sheet.max_row, min_col=1, max_col=Rpt_sheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000')) 
                    # adjust column width   
                    Rpt_sheet.column_dimensions['A'].width = 15
                    Rpt_sheet.column_dimensions['B'].width = 20
                    Rpt_sheet.column_dimensions['C'].width = 20  
                    Rpt_sheet.column_dimensions['D'].width = 15  
                    Rpt_sheet.column_dimensions['E'].width = 15

    for m in range (0,6):
        Rpt.worksheets[m].delete_rows(1,4)
        Rpt.worksheets[m].delete_cols(2)
        Rpt.worksheets[m].delete_cols(4)
        Rpt.worksheets[m].delete_cols(5)
        Rpt.worksheets[m].delete_cols(6,2)
    # fill first row with yellow color and apply bold font to tank inventory tabs for each site   
        for n in ['A1','B1','C1','D1','E1']:
            Rpt.worksheets[m][n].fill = PatternFill('solid', start_color='00FFFFCC')
            Rpt.worksheets[m][n].font = Font(bold=True)
    return Rpt

def compare_excel_sheets(file1, file2, old_date, new_date):
    # Load the two Excel files into dictionaries of DataFrames
    xls1 = pd.ExcelFile(file1)
    xls2 = pd.ExcelFile(file2)
    
    # Initialize a list to store the differences
    changes = []
    # Compare the first 6 sheets (tabs)
    for sheet_name in xls1.sheet_names[:6]:
        if sheet_name in xls2.sheet_names:
            # Load the data from both sheets
            df1 = pd.read_excel(xls1, sheet_name=sheet_name)
            df2 = pd.read_excel(xls2, sheet_name=sheet_name)
            
            # Add a 'Location' column to both DataFrames
            df1['Location'] = sheet_name
            df2['Location'] = sheet_name
            
            # Make sure the columns match
            if set(df1.columns) != set(df2.columns):
                print(f"Column names don't match in sheet {sheet_name}")
                continue
            
            # Iterate through the rows in df1 (old report)
            for index, row in df1.iterrows():
                # Find the corresponding row in the second sheet by matching 'Tank Name' and 'Location'
                matching_row = df2[(df2['Tank Name'] == row['Tank Name']) & (df2['Location'] == row['Location'])]
                if not matching_row.empty:
                    # If 'PRODUCT' is different, record the change
                    old_product = row['PRODUCT']
                    new_product = matching_row.iloc[0]['PRODUCT']
                    if old_product != new_product:
                        changes.append({
                            'Location': row['Location'],
                            'Tank Name': row['Tank Name'],
                            old_date: old_product,
                            new_date: new_product
                        })
            
            # Check for tanks in df2 (new report) that are not in df1 (old report)
            for index, row in df2.iterrows():
                matching_row = df1[(df1['Tank Name'] == row['Tank Name']) & (df1['Location'] == row['Location'])]
                if matching_row.empty:
                    # This tank is only in the new report
                    new_product = row['PRODUCT']
                    changes.append({
                        'Location': row['Location'],
                        'Tank Name': row['Tank Name'],
                        old_date: None,  # This tank did not exist in the old report
                        new_date: new_product
                    })

            # Check for tanks in df1 (old report) that are not in df2 (new report)
            for index, row in df1.iterrows():
                matching_row = df2[(df2['Tank Name'] == row['Tank Name']) & (df2['Location'] == row['Location'])]
                if matching_row.empty:
                    # This tank is only in the old report
                    old_product = row['PRODUCT']
                    changes.append({
                        'Location': row['Location'],
                        'Tank Name': row['Tank Name'],
                        old_date: old_product,  # This tank did not exist in the new report
                        new_date: None
                    })    
    return changes

def product_properties(synonyms_wb, changes_df, old_rpt, new_rpt):
    # Load the synonyms file
    synonyms_df = pd.read_excel(synonyms_wb, sheet_name='Chemicals 2024')

    # Convert the relevant product name columns to uppercase to ensure non case-insensitive merging
    synonyms_df['SYNONYM'] = synonyms_df['SYNONYM'].str.upper()
    changes_df[old_rpt] = changes_df[old_rpt].str.upper()
    changes_df[new_rpt] = changes_df[new_rpt].str.upper()
    
    # Merge LL/HL and OLD Status of Old Product 
    old_product_data = synonyms_df[['TERMINAL_NAME', 'SYNONYM', 'Service', 'OLD']]
    old_product_data.rename(columns={'TERMINAL_NAME': 'Location', 'SYNONYM': old_rpt, 'Service': 'Previous HL/LL Service','OLD':'Previous OLD Status'}, inplace=True)
    merged_df = pd.merge(changes_df, old_product_data, on=['Location', old_rpt], how='left')

    # Merge LL/HL and OLD Status of New Product
    new_product_data = synonyms_df[['TERMINAL_NAME', 'SYNONYM', 'Service', 'OLD']]
    new_product_data.rename(columns={'TERMINAL_NAME': 'Location', 'SYNONYM': new_rpt, 'Service': 'New HL/LL Service','OLD':'New OLD Status'}, inplace=True)
    merged_df = pd.merge(merged_df, new_product_data, on=['Location', new_rpt], how='left')
    
    # Rearrange the columns to insert the new ones behind the Old Product and New Product columns
    columns_order = ['Location', 'Tank Name', old_rpt, 'Previous HL/LL Service', 'Previous OLD Status', new_rpt, 'New HL/LL Service', 'New OLD Status']
    merged_df = merged_df[columns_order]
    return merged_df

def find_products_without_synonym(new_rpt_excel, synonyms_wb):
    # Seek synonyms_wb back to the beginning to allow re-reading
    if hasattr(synonyms_wb, 'seek'):
        synonyms_wb.seek(0)
    
    synonyms_df = pd.read_excel(synonyms_wb, sheet_name='Chemicals 2024')
    
    # Store clean uppercase (terminal, synonym) pairs
    synonyms_set = set()
    for _, row in synonyms_df.iterrows():
        term = str(row['TERMINAL_NAME']).strip().upper() if pd.notna(row['TERMINAL_NAME']) else ""
        syn = str(row['SYNONYM']).strip().upper() if pd.notna(row['SYNONYM']) else ""
        if term and syn:
            synonyms_set.add((term, syn))
            
    xls = pd.ExcelFile(new_rpt_excel)
    missing_synonyms = []
    
    # Check only the first 6 sheets (terminals)
    for sheet_name in xls.sheet_names[:6]:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        if 'PRODUCT' in df.columns:
            for _, row in df.iterrows():
                product = row['PRODUCT']
                if pd.notna(product):
                    prod_str = str(product).strip()
                    # Exclude empty products and common empty cell fillers
                    if prod_str and prod_str.upper() not in ['', 'EMPTY', 'OOS', 'OUT OF SERVICE', 'VACANT', 'NONE']:
                        loc_upper = str(sheet_name).strip().upper()
                        prod_upper = prod_str.upper()
                        if (loc_upper, prod_upper) not in synonyms_set:
                            missing_synonyms.append({
                                'Location': sheet_name,
                                'Product': prod_str
                            })
                            
    if missing_synonyms:
        return pd.DataFrame(missing_synonyms).drop_duplicates().reset_index(drop=True)
    else:
        return pd.DataFrame(columns=['Location', 'Product'])

#Streamlit app
st.set_page_config(layout="wide", page_title="Product Change Report")
st.write("## Gulf Air Product Change Report Generator")
# current and previous report dates
new_date = st.sidebar.date_input("Current Report Date (mm-dd-yyyy):",value='today',format='MM-DD-YYYY')
new_rpt = new_date.strftime("%m-%d-%Y") # MM-DD-YYYY
old_date = st.sidebar.date_input("Previous Report Date (mm-dd-yyyy):",value=new_date-timedelta(days=7),format='MM-DD-YYYY')
old_rpt = old_date.strftime("%m-%d-%Y") # MM-DD-YYYY
st.write("Upload current PAS.xlsx, GP.xlsx, GPWC.xlsx, JSTR.xlsx, KMET.xlsx and BOSTCO.xlsx")
new_rpt_wb = st.file_uploader("Choose `.xlsx` files", type="xlsx", accept_multiple_files=True)
st.write("Upload Previous Product Change Report")
old_rpt_wb = st.file_uploader(" Choose previous report `.xlsx` file", type="xlsx")
st.write("Upload Synonyms.xlsx")
synonyms_wb = st.file_uploader("Choose synonyms `.xlsx` file", type="xlsx")
ProductChange = 'Product Change '+ new_rpt + '.xlsx'
new_rpt_excel = new_rpt + '.xlsx'
if st.button("Generate Report"):
   # Load the Excel files
    merged_workbook = merge_inv_rpt(new_rpt_wb)
    merged_workbook.save(new_rpt_excel)
    # Compare the two reports
    changes = compare_excel_sheets(old_rpt_wb, new_rpt_excel, old_rpt, new_rpt)
    # Create a DataFrame from the changes
    changes_df = pd.DataFrame(changes)
    
    st.subheader("Detected Changes")
    # if there is no change, display a message and provide link to download the current week inventory
    if changes_df.empty:
        st.write("No product changes detected.")
        # Provide a button to download the result as an Excel file
        with open(new_rpt_excel, "rb") as file:
            st.download_button(label="Download Current Week Tank Inventory", data=file, file_name = ProductChange, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        # OOS and RTS are empty when changes_df is empty
        oos_display = pd.DataFrame(columns=['Location', 'Tank Name', 'Previous Product'])
        rts_display = pd.DataFrame(columns=['Location', 'Tank Name', 'Current Product'])
    else:
        # Get the properties of the changed products 
        merged_df = product_properties(synonyms_wb, changes_df, old_rpt, new_rpt)
        st.dataframe(merged_df)
        output_file = ProductChange
        merged_df.to_excel(output_file, index=False)
        with pd.ExcelWriter(new_rpt_excel, engine='openpyxl', mode='a') as writer:
            merged_df.to_excel(writer, sheet_name='Product Change '+new_rpt, index=False)
            # Apply styles to the 'Product Change' sheet
            product_change_sheet = writer.sheets['Product Change ' + new_rpt]
            for row in product_change_sheet.iter_rows(min_row=1, max_row=product_change_sheet.max_row, min_col=1, max_col=product_change_sheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'))
            # Adjust column widths
            column_widths = [15, 20, 20, 20, 20, 20, 20, 20]
            for i, width in enumerate(column_widths, start=1):
                product_change_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            # Fill first row with yellow color and apply bold font
            for cell in product_change_sheet["1:1"]:
                cell.fill = PatternFill('solid', start_color='00FFFFCC')
                cell.font = Font(bold=True)
        # Allow the user to download the report
        with open(new_rpt_excel, "rb") as f:
            st.download_button(
                label="Download Change Report",
                data=f,
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # OOS: Existed in old report (old_rpt is not null) but not in new report (new_rpt is null)
        oos_df = merged_df[merged_df[new_rpt].isna() | (merged_df[new_rpt] == '')]
        oos_display = oos_df[['Location', 'Tank Name', old_rpt]].rename(columns={old_rpt: 'Previous Product'})
        
        # RTS: Not in old report (old_rpt is null) but in new report (new_rpt is not null)
        rts_df = merged_df[merged_df[old_rpt].isna() | (merged_df[old_rpt] == '')]
        rts_display = rts_df[['Location', 'Tank Name', new_rpt]].rename(columns={new_rpt: 'Current Product'})

    # Get products without synonyms from current week's sheets
    missing_syn_df = find_products_without_synonym(new_rpt_excel, synonyms_wb)

    # Display three tables as summary reports
    st.markdown("---")
    st.markdown("<h2 style='text-align: center; color: #1E3A8A;'>📊 Summary Reports</h2>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div style='background-color: #FEF2F2; border-left: 5px solid #EF4444; padding: 15px; border-radius: 4px; margin-bottom: 10px;'>
            <h3 style='margin: 0; color: #991B1B;'>🚫 Tanks Out of Service</h3>
            <p style='margin: 5px 0 0 0; color: #7F1D1D; font-size: 14px;'>Tanks in previous report but missing in current report</p>
        </div>
        """, unsafe_allow_html=True)
        if not oos_display.empty:
            st.dataframe(oos_display, use_container_width=True, hide_index=True)
        else:
            st.success("No tanks moved Out of Service.")
            
    with col2:
        st.markdown("""
        <div style='background-color: #F0FDF4; border-left: 5px solid #22C55E; padding: 15px; border-radius: 4px; margin-bottom: 10px;'>
            <h3 style='margin: 0; color: #166534;'>🔄 Tanks Returned to Service</h3>
            <p style='margin: 5px 0 0 0; color: #14532D; font-size: 14px;'>Tanks not in previous report but in current report</p>
        </div>
        """, unsafe_allow_html=True)
        if not rts_display.empty:
            st.dataframe(rts_display, use_container_width=True, hide_index=True)
        else:
            st.success("No tanks returned to Service.")
            
    with col3:
        st.markdown("""
        <div style='background-color: #FFF7ED; border-left: 5px solid #F97316; padding: 15px; border-radius: 4px; margin-bottom: 10px;'>
            <h3 style='margin: 0; color: #9A3412;'>⚠️ New Products without Synonym</h3>
            <p style='margin: 5px 0 0 0; color: #7C2D12; font-size: 14px;'>Products in current report not in Synonyms.xlsx</p>
        </div>
        """, unsafe_allow_html=True)
        if not missing_syn_df.empty:
            st.dataframe(missing_syn_df, use_container_width=True, hide_index=True)
        else:
            st.success("All products have valid synonyms!")
